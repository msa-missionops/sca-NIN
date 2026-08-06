"""Tests for reference/tag data loading (Phase 1F)."""

import openpyxl
import pytest
from openpyxl.worksheet.table import Table

from nin_pipeline.reference_data import (
    active_plant,
    load_reference_data,
    load_reference_data_from_workbook,
)


def write_reference_data(folder):
    folder.mkdir(parents=True, exist_ok=True)
    (folder / "region.csv").write_text("Plant,Region\nUS01,Northeast\n")
    (folder / "top60.csv").write_text(
        "plant_material_key,plant,material,top_60_flag\nUS01-123,US01,123,top60\n"
    )
    (folder / "sourceplant.csv").write_text(
        "source_key,desc,source_plant\n40,Subcontract,US02\n"
    )
    (folder / "rec_req_type.csv").write_text(
        "type,negative\nVJ,true\nPP,false\nU1,1\nWB,0\nVC,yes\nVG,no\n"
    )
    (folder / "plant_evaluation.csv").write_text("Plant\n us01 \n")


def _add_table(ws, display_name, rows, start_row=1):
    """Write `rows` (a list of row-lists, first row is the header) at
    `start_row` and register them as a named Excel Table, mirroring how
    business users maintain these tables as ListObjects in the real
    workbook (multiple tables can share a worksheet, stacked or side by
    side)."""
    for offset, row in enumerate(rows):
        for col, value in enumerate(row, start=1):
            ws.cell(row=start_row + offset, column=col, value=value)
    n_cols = len(rows[0])
    last_col = openpyxl.utils.get_column_letter(n_cols)
    end_row = start_row + len(rows) - 1
    ws.add_table(
        Table(displayName=display_name, ref=f"A{start_row}:{last_col}{end_row}")
    )


def write_reference_workbook(path):
    """Build a workbook with the same 5 named Excel Tables (across multiple
    worksheets, out of order, mixed case) that production's tables use, to
    confirm lookup is by table name (not sheet name/position)."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Tags1"
    _add_table(ws1, "tbl_Tag_Region", [["Plant", "Region"], ["US01", "Northeast"]])

    ws2 = wb.create_sheet("Tags2")
    _add_table(
        ws2,
        "tbl_Tag_Top60",
        [
            ["plant_material_key", "plant", "material", "top_60_flag"],
            ["US01-123", "US01", "123", "top 60"],
        ],
        start_row=1,
    )
    _add_table(
        ws2,
        "tbl_tag_sourceplant",
        [["source_key", "desc", "source_plant"], ["40", "Subcontract", "US02"]],
        start_row=5,
    )

    ws3 = wb.create_sheet("Config")
    _add_table(
        ws3,
        "rec_req_type",
        [["type", "negative"], ["VJ", "TRUE"], ["PP", "FALSE"]],
        start_row=1,
    )
    _add_table(ws3, "plant_evaluation", [["Plant"], [" us01 "]], start_row=6)
    wb.save(path)


def test_load_reference_data_parses_negative_flags_case_insensitively(tmp_path):
    folder = tmp_path / "reference_data"
    write_reference_data(folder)

    ref = load_reference_data(folder)

    rec = ref.rec_req_type.set_index("type")["negative"]
    assert bool(rec["VJ"]) is True
    assert bool(rec["PP"]) is False
    assert bool(rec["U1"]) is True
    assert bool(rec["WB"]) is False
    assert bool(rec["VC"]) is True
    assert bool(rec["VG"]) is False


def test_active_plant_trims_and_upper_cases(tmp_path):
    folder = tmp_path / "reference_data"
    write_reference_data(folder)
    ref = load_reference_data(folder)

    assert active_plant(ref) == "US01"


def test_load_reference_data_raises_when_file_missing(tmp_path):
    folder = tmp_path / "reference_data"
    folder.mkdir()

    with pytest.raises(FileNotFoundError):
        load_reference_data(folder)


def test_load_reference_data_from_workbook_reads_named_tables_across_sheets(tmp_path):
    workbook_path = tmp_path / "reference_data.xlsx"
    write_reference_workbook(workbook_path)

    ref = load_reference_data_from_workbook(workbook_path)

    assert ref.region.iloc[0]["Region"] == "Northeast"
    assert ref.top60.iloc[0]["top_60_flag"] == "top 60"
    assert ref.sourceplant.iloc[0]["source_plant"] == "US02"
    assert active_plant(ref) == "US01"

    rec = ref.rec_req_type.set_index("type")["negative"]
    assert bool(rec["VJ"]) is True
    assert bool(rec["PP"]) is False


def test_load_reference_data_from_workbook_raises_when_table_missing(tmp_path):
    workbook_path = tmp_path / "reference_data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    _add_table(ws, "tbl_Tag_Region", [["Plant", "Region"], ["US01", "Northeast"]])
    wb.save(workbook_path)

    with pytest.raises(ValueError, match="tbl_Tag_Top60"):
        load_reference_data_from_workbook(workbook_path)
