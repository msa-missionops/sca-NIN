"""End-to-end integration test for the pipeline orchestration (Phase 1F)."""

import openpyxl
import pandas as pd
from openpyxl.worksheet.table import Table

from nin_pipeline.config import PipelineConfig, load_config
from nin_pipeline.pipeline import run_pipeline

PRDPL3_HEADER = [
    "Plnt",
    "Material",
    "Material Description",
    "BUn",
    "Product hierarchy",
    "Basic material",
    "Basic material 2",
    "MRP Typ",
    "MRP Controller",
    "PuRGrp",
    "DelFlag",
    "MtlSt_XPlt",
    "MtlSt_Plt",
    "X-Chn Mtl St",
    "SG",
    "ProcType",
    "Spec Proc",
    "ABC",
    "MTyp",
    "IPT",
    "PDT",
    "GRT",
    "TRLT",
    "Planning time fence",
    "LotSize",
    "Min. Lot Sze",
    "Rounding val.",
    "Max. Lot Size",
    "Fix. lot size",
    "BUn_1",
    "Safety Stock",
    "Reorder Point",
    "Threshold Qty",
    "AvaiChk",
    "Consumption mode",
    "Bwd cons. per.",
    "Fwd cons.period",
    "BkFlush",
    "PSloc",
    "Language",
    "Tot Valuated Stk",
    "Total Value",
    "Standard price",
    "Price Un",
    "SLife",
    "Propos.SA",
    "ESLoc",
    "R. Profile",
    "ProdS",
]


def _pipe_row(header, overrides):
    values = {name: "" for name in header}
    values.update(overrides)
    return "|" + "|".join(values[name] for name in header) + "|"


def write_prdpl3_export(path, data_rows):
    lines = [f"BANNER LINE {i}" for i in range(6)]
    lines.append(_pipe_row(PRDPL3_HEADER, dict(zip(PRDPL3_HEADER, PRDPL3_HEADER))))
    lines.append(_pipe_row(PRDPL3_HEADER, {}))
    lines.extend(_pipe_row(PRDPL3_HEADER, row) for row in data_rows)
    path.write_text("\n".join(lines), encoding="cp1252")


MB5T_HEADER = [
    "Material",
    "Material Description",
    "Quantity",
    "Plnt",
    "Name 1",
    "Pur. Doc.",
    "Item",
    "SPlt",
    "S",
    "BUn",
    "Amount in LC",
    "Crcy",
    "Quantity_1",
    "OUn",
    "Net Value",
    "Crcy_2",
]


def write_mb5t_export(path, data_rows):
    lines = [f"BANNER LINE {i}" for i in range(3)]
    lines.append(_pipe_row(MB5T_HEADER, dict(zip(MB5T_HEADER, MB5T_HEADER))))
    lines.append(_pipe_row(MB5T_HEADER, {}))
    lines.extend(_pipe_row(MB5T_HEADER, row) for row in data_rows)
    path.write_text("\n".join(lines), encoding="cp1252")


DOH_HEADER = [
    "Plnt",
    "Material",
    "El",
    "Customer Request Date",
    "Rec./reqd.qty",
    "BUn",
]


def write_doh_export(path, data_rows):
    lines = ["Banner text 0", "Banner text 1", ""]
    lines.append("\t".join(DOH_HEADER))
    for row in data_rows:
        lines.append("\t".join(row.get(col, "") for col in DOH_HEADER))
    path.write_text("\n".join(lines), encoding="cp1252")


REC_HEADER = [
    "Plnt",
    "Material",
    "El",
    "Customer Request Date",
    "Rec./reqd.qty",
    "BUn",
]


def write_rec_export(path, data_rows):
    lines = ["Banner text 0", "Banner text 1", ""]
    lines.append("\t".join(REC_HEADER))
    for row in data_rows:
        lines.append("\t".join(row.get(col, "") for col in REC_HEADER))
    path.write_text("\n".join(lines), encoding="cp1252")


def write_reference_data(folder):
    folder.mkdir(parents=True, exist_ok=True)
    (folder / "region.csv").write_text("Plant,Region\nUS01,Northeast\n")
    (folder / "top60.csv").write_text("plant_material_key,plant,material,top_60_flag\n")
    (folder / "sourceplant.csv").write_text("source_key,desc,source_plant\n")
    (folder / "rec_req_type.csv").write_text("type,negative\n")
    (folder / "plant_evaluation.csv").write_text("Plant\nUS01\n")


def write_reference_workbook(path):
    """Build the same reference data as `write_reference_data`, but as a
    single Excel workbook with named Tables -- the format business users
    actually maintain directly in Excel (see
    `nin_pipeline.reference_data.load_reference_data_from_workbook`)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tags"

    def add_table(name, rows, start_row):
        for offset, row in enumerate(rows):
            for col, value in enumerate(row, start=1):
                ws.cell(row=start_row + offset, column=col, value=value)
        last_col = openpyxl.utils.get_column_letter(len(rows[0]))
        end_row = start_row + len(rows) - 1
        ws.add_table(Table(displayName=name, ref=f"A{start_row}:{last_col}{end_row}"))

    add_table("tbl_Tag_Region", [["Plant", "Region"], ["US01", "Northeast"]], 1)
    add_table(
        "tbl_Tag_Top60",
        [["plant_material_key", "plant", "material", "top_60_flag"]],
        4,
    )
    add_table("tbl_tag_sourceplant", [["source_key", "desc", "source_plant"]], 7)
    add_table("rec_req_type", [["type", "negative"]], 10)
    add_table("plant_evaluation", [["Plant"], ["US01"]], 13)
    wb.save(path)


def make_config(tmp_path) -> PipelineConfig:
    config_path = tmp_path / "settings.yaml"
    config_path.write_text(
        f"""
paths:
  prdpl3_folder: "{tmp_path / 'prdpl3'}"
  mrp_rec_folder: "{tmp_path / 'mrp_rec'}"
  mrp_doh_folder: "{tmp_path / 'mrp_doh'}"
  mb5t_folder: "{tmp_path / 'mb5t'}"
  reference_data_folder: "{tmp_path / 'reference_data'}"
  output_folder: "{tmp_path / 'output'}"
  run_folder: "{tmp_path / 'runs'}"
  log_folder: "{tmp_path / 'logs'}"
""",
        encoding="utf-8",
    )
    return load_config(config_path)


def make_config_with_workbook(tmp_path) -> PipelineConfig:
    config_path = tmp_path / "settings.yaml"
    config_path.write_text(
        f"""
paths:
  prdpl3_folder: "{tmp_path / 'prdpl3'}"
  mrp_rec_folder: "{tmp_path / 'mrp_rec'}"
  mrp_doh_folder: "{tmp_path / 'mrp_doh'}"
  mb5t_folder: "{tmp_path / 'mb5t'}"
  reference_data_workbook: "{tmp_path / 'reference_data.xlsx'}"
  output_folder: "{tmp_path / 'output'}"
  run_folder: "{tmp_path / 'runs'}"
  log_folder: "{tmp_path / 'logs'}"
""",
        encoding="utf-8",
    )
    return load_config(config_path)


def make_config_with_active_plants(tmp_path) -> PipelineConfig:
    config_path = tmp_path / "settings.yaml"
    config_path.write_text(
        f"""
paths:
  prdpl3_folder: "{tmp_path / 'prdpl3'}"
  mrp_rec_folder: "{tmp_path / 'mrp_rec'}"
  mrp_doh_folder: "{tmp_path / 'mrp_doh'}"
  mb5t_folder: "{tmp_path / 'mb5t'}"
  reference_data_folder: "{tmp_path / 'reference_data'}"
  active_plants_folder: "{tmp_path / 'active_plants'}"
  output_folder: "{tmp_path / 'output'}"
  run_folder: "{tmp_path / 'runs'}"
  log_folder: "{tmp_path / 'logs'}"
""",
        encoding="utf-8",
    )
    return load_config(config_path)


def make_config_with_csv_mrp_extension(tmp_path) -> PipelineConfig:
    config_path = tmp_path / "settings.yaml"
    config_path.write_text(
        f"""
paths:
  prdpl3_folder: "{tmp_path / 'prdpl3'}"
  mrp_rec_folder: "{tmp_path / 'mrp_rec'}"
  mrp_doh_folder: "{tmp_path / 'mrp_doh'}"
  mb5t_folder: "{tmp_path / 'mb5t'}"
  reference_data_folder: "{tmp_path / 'reference_data'}"
  output_folder: "{tmp_path / 'output'}"
  run_folder: "{tmp_path / 'runs'}"
  log_folder: "{tmp_path / 'logs'}"

file_selection:
  mrp_doh_file_extension: csv
  mrp_rec_file_extension: csv
""",
        encoding="utf-8",
    )
    return load_config(config_path)


def test_run_pipeline_assembles_base_table_end_to_end(tmp_path):
    prdpl3_folder = tmp_path / "prdpl3"
    prdpl3_folder.mkdir()
    write_prdpl3_export(
        prdpl3_folder / "PRDPL3_export.txt",
        [
            {
                "Plnt": "US01",
                "Material": "000123",
                "Product hierarchy": "00020ABC",
                "DelFlag": "",
                "Tot Valuated Stk": "200",
                "Total Value": "5000",
                "Safety Stock": "10",
            }
        ],
    )

    mb5t_folder = tmp_path / "mb5t"
    mb5t_folder.mkdir()
    write_mb5t_export(
        mb5t_folder / "MB5T_export.txt",
        [{"Plnt": "US01", "Material": "000123", "Quantity": "7"}],
    )

    doh_run_folder = tmp_path / "mrp_doh" / "20260201_run"
    doh_run_folder.mkdir(parents=True)
    write_doh_export(
        doh_run_folder / "MM_MRP_ELEMENTS_DOH_20260201_US01_x.txt",
        [
            {
                "Plnt": "US01",
                "Material": "000123",
                "El": "VJ",
                "Customer Request Date": "2026-12-06",
                "Rec./reqd.qty": "30",
                "BUn": "EA",
            }
        ],
    )

    rec_run_folder = tmp_path / "mrp_rec" / "20260201_run"
    rec_run_folder.mkdir(parents=True)
    write_rec_export(
        rec_run_folder / "MM_MRP_ELEMENTS_REC_20260201_US01_x.txt",
        [
            {
                "Plnt": "US01",
                "Material": "000123",
                "El": "VJ",
                "Customer Request Date": "2026-12-04",  # Friday -> week 1
                "Rec./reqd.qty": "10",
                "BUn": "EA",
            },
            {
                "Plnt": "US01",
                "Material": "000123",
                "El": "VJ",
                "Customer Request Date": "2026-12-11",  # next Friday -> week 2
                "Rec./reqd.qty": "5",
                "BUn": "EA",
            },
        ],
    )

    write_reference_data(tmp_path / "reference_data")

    config = make_config(tmp_path)
    result = run_pipeline(config, run_id="20260201_000000")

    assert len(result.base_table) == 1
    row = result.base_table.iloc[0]
    assert row["plant_material_key"] == "US01-123"
    assert row["Quantity in Transit"] == 7
    assert row["VJ"] == 30
    # Available Stock = max(0, 200 - (30+0+0+0)) = 170
    assert row["Available Stock"] == 170
    # BOBL processing is deferred for now (Open Decision #10) -- placeholders
    # are always null.
    assert pd.isna(row["Backorder Actual"])
    assert pd.isna(row["Backlog Qnty"])

    # Weekly REC forecast transpose: week 1 = earliest Week Ending (10),
    # week 2 = next (5), remaining weeks default to 0, Total Forecast (Qty)
    # = sum of week 1..27 only.
    assert row["week 1"] == 10
    assert row["week 2"] == 5
    assert row["week 3"] == 0
    assert row["Total Forecast (Qty)"] == 15

    assert result.manifest_path.exists()
    manifest = pd.read_json(result.manifest_path, typ="series")
    assert manifest["run_id"] == "20260201_000000"
    assert manifest["row_counts"]["nin_base_table"] == 1
    assert manifest["row_counts"]["plants"]["US01"]["mrp_rec_raw"] == 2

    output_folder = tmp_path / "output"
    assert (output_folder / "nin_base_table.csv").exists()
    assert (output_folder / "nin_base_table.parquet").exists()
    assert (output_folder / "nin_base_table.xlsx").exists()


def test_run_pipeline_respects_configured_mrp_file_extension(tmp_path):
    """Live SAP MRP_ELEMENTS_DOH/REC exports may be written as .csv rather
    than the .txt default (see docs/nin_data_contracts.md); the
    file_selection.mrp_doh_file_extension/mrp_rec_file_extension settings
    let the file-discovery step be pointed at whichever extension is
    actually produced."""
    prdpl3_folder = tmp_path / "prdpl3"
    prdpl3_folder.mkdir()
    write_prdpl3_export(
        prdpl3_folder / "PRDPL3_export.txt",
        [
            {
                "Plnt": "US01",
                "Material": "000123",
                "Product hierarchy": "00020ABC",
                "DelFlag": "",
                "Tot Valuated Stk": "200",
                "Total Value": "5000",
                "Safety Stock": "10",
            }
        ],
    )

    mb5t_folder = tmp_path / "mb5t"
    mb5t_folder.mkdir()
    write_mb5t_export(
        mb5t_folder / "MB5T_export.txt",
        [{"Plnt": "US01", "Material": "000123", "Quantity": "7"}],
    )

    doh_run_folder = tmp_path / "mrp_doh" / "20260201_run"
    doh_run_folder.mkdir(parents=True)
    write_doh_export(
        doh_run_folder / "MM_MRP_ELEMENTS_DOH_20260201_US01_x.csv",
        [
            {
                "Plnt": "US01",
                "Material": "000123",
                "El": "VJ",
                "Customer Request Date": "2026-12-06",
                "Rec./reqd.qty": "30",
                "BUn": "EA",
            }
        ],
    )

    rec_run_folder = tmp_path / "mrp_rec" / "20260201_run"
    rec_run_folder.mkdir(parents=True)
    write_rec_export(
        rec_run_folder / "MM_MRP_ELEMENTS_REC_20260201_US01_x.csv",
        [
            {
                "Plnt": "US01",
                "Material": "000123",
                "El": "VJ",
                "Customer Request Date": "2026-12-04",
                "Rec./reqd.qty": "10",
                "BUn": "EA",
            }
        ],
    )

    write_reference_data(tmp_path / "reference_data")

    config = make_config_with_csv_mrp_extension(tmp_path)
    result = run_pipeline(config, run_id="20260201_000000")

    assert len(result.base_table) == 1
    row = result.base_table.iloc[0]
    assert row["plant_material_key"] == "US01-123"
    assert row["VJ"] == 30
    assert row["week 1"] == 10


def test_run_pipeline_accepts_reference_data_workbook_instead_of_csv_folder(tmp_path):
    """Business users maintain the tag tables as named Excel Tables inside
    a single workbook, not as loose CSVs -- confirm the pipeline can be
    pointed directly at that workbook (see
    `nin_pipeline.reference_data.load_reference_data_from_workbook`),
    requiring no manual CSV export step."""
    prdpl3_folder = tmp_path / "prdpl3"
    prdpl3_folder.mkdir()
    write_prdpl3_export(
        prdpl3_folder / "PRDPL3_export.txt",
        [
            {
                "Plnt": "US01",
                "Material": "000123",
                "Product hierarchy": "00020ABC",
                "DelFlag": "",
                "Tot Valuated Stk": "200",
                "Total Value": "5000",
                "Safety Stock": "10",
            }
        ],
    )

    mb5t_folder = tmp_path / "mb5t"
    mb5t_folder.mkdir()
    write_mb5t_export(
        mb5t_folder / "MB5T_export.txt",
        [{"Plnt": "US01", "Material": "000123", "Quantity": "7"}],
    )

    doh_run_folder = tmp_path / "mrp_doh" / "20260201_run"
    doh_run_folder.mkdir(parents=True)
    write_doh_export(
        doh_run_folder / "MM_MRP_ELEMENTS_DOH_20260201_US01_x.txt",
        [
            {
                "Plnt": "US01",
                "Material": "000123",
                "El": "VJ",
                "Customer Request Date": "2026-12-06",
                "Rec./reqd.qty": "30",
                "BUn": "EA",
            }
        ],
    )

    rec_run_folder = tmp_path / "mrp_rec" / "20260201_run"
    rec_run_folder.mkdir(parents=True)
    write_rec_export(
        rec_run_folder / "MM_MRP_ELEMENTS_REC_20260201_US01_x.txt",
        [
            {
                "Plnt": "US01",
                "Material": "000123",
                "El": "VJ",
                "Customer Request Date": "2026-12-04",
                "Rec./reqd.qty": "10",
                "BUn": "EA",
            }
        ],
    )

    write_reference_workbook(tmp_path / "reference_data.xlsx")

    config = make_config_with_workbook(tmp_path)
    result = run_pipeline(config, run_id="20260201_000000")

    assert len(result.base_table) == 1
    row = result.base_table.iloc[0]
    assert row["plant_material_key"] == "US01-123"
    assert row["Quantity in Transit"] == 7
    assert row["VJ"] == 30
    assert row["week 1"] == 10


def test_run_pipeline_runs_for_every_plant_in_active_plants_folder(tmp_path):
    """When `paths.active_plants_folder` is configured, the pipeline runs
    once per plant listed in the latest file there (a headerless,
    single-column CSV -- see `nin_pipeline.reference_data.load_active_plants`)
    and concatenates every plant's base table into one combined result,
    instead of using the single plant in `plant_evaluation.csv`."""
    prdpl3_folder = tmp_path / "prdpl3"
    prdpl3_folder.mkdir()
    write_prdpl3_export(
        prdpl3_folder / "PRDPL3_export.txt",
        [
            {
                "Plnt": "US01",
                "Material": "000123",
                "Product hierarchy": "00020ABC",
                "DelFlag": "",
                "Tot Valuated Stk": "200",
                "Total Value": "5000",
                "Safety Stock": "10",
            },
            {
                "Plnt": "US02",
                "Material": "000456",
                "Product hierarchy": "00020DEF",
                "DelFlag": "",
                "Tot Valuated Stk": "300",
                "Total Value": "6000",
                "Safety Stock": "20",
            },
        ],
    )

    mb5t_folder = tmp_path / "mb5t"
    mb5t_folder.mkdir()
    write_mb5t_export(
        mb5t_folder / "MB5T_export.txt",
        [
            {"Plnt": "US01", "Material": "000123", "Quantity": "7"},
            {"Plnt": "US02", "Material": "000456", "Quantity": "9"},
        ],
    )

    doh_run_folder = tmp_path / "mrp_doh" / "20260201_run"
    doh_run_folder.mkdir(parents=True)
    write_doh_export(
        doh_run_folder / "MM_MRP_ELEMENTS_DOH_20260201_US01_x.txt",
        [
            {
                "Plnt": "US01",
                "Material": "000123",
                "El": "VJ",
                "Customer Request Date": "2026-12-06",
                "Rec./reqd.qty": "30",
                "BUn": "EA",
            }
        ],
    )
    write_doh_export(
        doh_run_folder / "MM_MRP_ELEMENTS_DOH_20260201_US02_x.txt",
        [
            {
                "Plnt": "US02",
                "Material": "000456",
                "El": "VJ",
                "Customer Request Date": "2026-12-06",
                "Rec./reqd.qty": "40",
                "BUn": "EA",
            }
        ],
    )

    rec_run_folder = tmp_path / "mrp_rec" / "20260201_run"
    rec_run_folder.mkdir(parents=True)
    write_rec_export(
        rec_run_folder / "MM_MRP_ELEMENTS_REC_20260201_US01_x.txt",
        [
            {
                "Plnt": "US01",
                "Material": "000123",
                "El": "VJ",
                "Customer Request Date": "2026-12-04",
                "Rec./reqd.qty": "10",
                "BUn": "EA",
            }
        ],
    )
    write_rec_export(
        rec_run_folder / "MM_MRP_ELEMENTS_REC_20260201_US02_x.txt",
        [
            {
                "Plnt": "US02",
                "Material": "000456",
                "El": "VJ",
                "Customer Request Date": "2026-12-04",
                "Rec./reqd.qty": "20",
                "BUn": "EA",
            }
        ],
    )

    write_reference_data(tmp_path / "reference_data")

    active_plants_folder = tmp_path / "active_plants"
    active_plants_folder.mkdir()
    (active_plants_folder / "active_plants.csv").write_text("US01\nUS02\n")

    config = make_config_with_active_plants(tmp_path)
    result = run_pipeline(config, run_id="20260201_000000")

    assert len(result.base_table) == 2
    rows = result.base_table.set_index("plant_material_key")
    assert set(rows.index) == {"US01-123", "US02-456"}
    assert rows.loc["US01-123", "Quantity in Transit"] == 7
    assert rows.loc["US02-456", "Quantity in Transit"] == 9
    assert rows.loc["US01-123", "VJ"] == 30
    assert rows.loc["US02-456", "VJ"] == 40

    manifest = pd.read_json(result.manifest_path, typ="series")
    assert manifest["row_counts"]["nin_base_table"] == 2
    assert set(manifest["row_counts"]["plants"].keys()) == {"US01", "US02"}
    assert manifest["sources"]["plants"]["US01"]["mrp_rec"].endswith(
        "MM_MRP_ELEMENTS_REC_20260201_US01_x.txt"
    )
    assert (
        str(active_plants_folder / "active_plants.csv")
        == manifest["sources"]["active_plants_file"]
    )
